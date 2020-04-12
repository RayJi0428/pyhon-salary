#https://openpyxl.readthedocs.io/en/stable/tutorial.html#
import openpyxl  # excel
import os  # 檔案
import codecs  # 檔案
import json  # JSON
import subprocess  # 子進程

import smtplib  # email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import sys
# --------------------------------------------------------------------------------
# encrypt
# https://stackoverflow.com/questions/39509741/python-or-libreoffice-save-xlsx-file-encrypted-with-password
# https://stackoverflow.com/questions/36122496/password-protecting-excel-file-using-python

# email
# https://www.runoob.com/python/python-email.html
# SMTP(Simple Mail Transfer Protocol)
# --------------------------------------------------------------------------------


def set_password(excel_file_path, pw):

    from pathlib import Path

    excel_file_path = Path(excel_file_path)

    vbs_script = \
        f"""' Save with password required upon opening

    Set excel_object = CreateObject("Excel.Application")
    Set workbook = excel_object.Workbooks.Open("{excel_file_path}")

    excel_object.DisplayAlerts = False
    excel_object.Visible = False

    workbook.SaveAs "{excel_file_path}",, "{pw}"

    excel_object.Application.Quit
    """

    # write
    vbs_script_path = excel_file_path.parent.joinpath(
        excel_file_path.stem + "set_pw.vbs")
    with open(vbs_script_path, "w") as file:
        file.write(vbs_script)

    # execute
    subprocess.call(['cscript.exe', str(vbs_script_path)])

    # remove
    vbs_script_path.unlink()

    return None

# --------------------------------------------------------------------------------


def sendExcelByMail():
    try:
        # smtpObj = smtplib.SMTP('smtp.gmail.com', 587)  # TLS
        smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', 465)  # SSL
    except:
        print('呼叫SMTP失敗!')
    response = smtpObj.ehlo()  # 對SMTP打招呼

    if response[0] != 250:
        print('SMTP ehlo失敗!')  # 返回的tuple第一項是250表示成功
    # smtpObj.starttls()  # 啟動TLS加密(SSL可省略此步驟)

    sender = ""
    password = ""
    receivers = [""]
    mail_subject = "TestSubject"
    # 低安全性應用程式存取權
    # https://support.google.com/mail/?p=BadCredentials
    response = smtpObj.login(sender, password)

    msg = MIMEMultipart()
    msg["From"] = sender  # 發件人
    msg["To"] = ";".join(receivers)  # 收件人
    msg["Subject"] = mail_subject   # 郵件標題
    # 構造附件
    #att = MIMEText(aaa, "plain", "utf-8")
    xlsx = MIMEBase(
        'application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment = open("test1.xlsx", "rb")
    xlsx.set_payload(attachment.read())
    encoders.encode_base64(xlsx)
    xlsx.add_header('Content-Disposition', 'attachment;filename="test1.xlsx"')
    msg.attach(xlsx)

    smtpObj.sendmail(sender, receivers, msg.as_string())

# --------------------------------------------------------------------------------


argvs = sys.argv[1:]
print(argvs)
code = argvs[0]
name = argvs[1]
pwd = argvs[2]
email = argvs[3]
target_row = int(argvs[4])
print("main2:", code, name, pwd, email)
input_file = os.path.join(os.getcwd(), 'test.xlsx')
wb2 = openpyxl.load_workbook(input_file)
ws2 = wb2.worksheets[0]
for ws2_ri in range(30, 2, -1):
    if ws2_ri != target_row:
        ws2.delete_rows(ws2_ri)  # 1-base
# wb2 = openpyxl.Workbook()
# ws2 = wb2.active
# ws2.append(ws[1])
# ws2.append(ws[2])
# ws2.append(ws[ri])
# ws.delete_rows(3)  # 1-base
output_file = os.path.join(os.getcwd(), code + '.xlsx')
wb2.save(output_file)
set_password(output_file, pwd)
