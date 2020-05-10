#https://openpyxl.readthedocs.io/en/stable/tutorial.html#
import shutil
import openpyxl  # excel
import os  # 檔案
import codecs  # 檔案
import json  # JSON
import subprocess  # 子進程
import sys
import zipfile
import smtplib  # email
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
# --------------------------------------------------------------------------------
# encrypt
# https://stackoverflow.com/questions/39509741/python-or-libreoffice-save-xlsx-file-encrypted-with-password
# https://stackoverflow.com/questions/36122496/password-protecting-excel-file-using-python

# email
# https://www.runoob.com/python/python-email.html
# SMTP(Simple Mail Transfer Protocol)
# --------------------------------------------------------------------------------

# zip加密


def set_password2(excel_file_path, pw):
    zipPath = excel_file_path.replace('xlsx', 'zip')
    newZip = zipfile.ZipFile(zipPath, 'w')
    newZip.write(excel_file_path, compress_type=zipfile.ZIP_DEFLATED)
    newZip.setpassword(bytes(pw, 'ascii'))
    newZip.close()

# vbs加密


def set_password(excel_file_path, pw):

    from pathlib import Path

    excel_file_path = Path(excel_file_path)

    vbs_script = \
        f"""' Save with password required upon opening

    Set excel_object = CreateObject("Excel.Application")
    excel_object.DisplayAlerts = False
    excel_object.Visible = False

    Set workbook = excel_object.Workbooks.Open("{excel_file_path}")

    workbook.SaveAs "{excel_file_path}",, "{pw}"

    excel_object.Application.Quit
    """

    # write
    vbs_script_path = excel_file_path.parent.joinpath("set_pw.vbs")
    with open(vbs_script_path, "w") as file:
        file.write(vbs_script)

    # execute
    subprocess.call(['cscript.exe', str(vbs_script_path)])

    # remove
    vbs_script_path.unlink()

    return None

# --------------------------------------------------------------------------------


def sendExcelByMail(subject, file, receiver):
    try:
        # smtpObj = smtplib.SMTP('smtp.gmail.com', 587)  # TLS
        smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', 465)  # SSL
    except:
        print('呼叫SMTP失敗!')
    response = smtpObj.ehlo()  # 對SMTP打招呼

    if response[0] != 250:
        print('SMTP ehlo失敗!')  # 返回的tuple第一項是250表示成功
    # smtpObj.starttls()  # 啟動TLS加密(SSL可省略此步驟)

    sender = sender_account
    password = sender_pwd
    receivers = [receiver]
    mail_subject = subject
    # 低安全性應用程式存取權
    # https://support.google.com/mail/?p=BadCredentials
    response = smtpObj.login(sender, password)

    msg = MIMEMultipart()
    msg["From"] = sender  # 發件人
    msg["To"] = ";".join(receivers)  # 收件人
    msg["Subject"] = mail_subject   # 郵件標題
    # 構造附件
    xlsx = MIMEBase(
        'application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment = open(file, "rb")
    xlsx.set_payload(attachment.read())
    encoders.encode_base64(xlsx)
    attachment_str = 'attachment;filename="#"'.replace(
        "#", os.path.basename(file))
    xlsx.add_header('Content-Disposition', attachment_str)
    msg.attach(xlsx)

    smtpObj.sendmail(sender, receivers, msg.as_string())

# --------------------------------------------------------------------------------


def loadJSON(filename):
    json_file = codecs.open(filename, 'r', encoding='utf-8')
    try:
        json_data = json.loads(json_file.read())
    except:
        input(filename + ' 解析失敗')

    json_file.close()
    return json_data

# --------------------------------------------------------------------------------


# 讀取人員設定表
config = loadJSON('config.json')
sender_account = config['Sender']['Account']
sender_pwd = config['Sender']['pwd']

tmp_dir = os.path.join(os.getcwd(), 'tmp')
if os.path.isdir(tmp_dir):
    shutil.rmtree(tmp_dir)

# 讀取excel
# input_file = os.path.join(os.getcwd(), 'test.xlsx')

if len(sys.argv) == 1:
    input_file = input('請將薪資表拖曳至視窗內:')
else:
    input_file = sys.argv[1]

print('開始拆分薪資表...' + input_file)

wb = openpyxl.load_workbook(input_file)
ws = wb.worksheets[0]  # 取第一張表

# 建立暫存資料夾--------------------------------------------------------------------------------

os.mkdir(tmp_dir)

tmp_files = []

# 薪資拆分--------------------------------------------------------------------------------

# 資料起始列為3,1:日期列, 2:標題列
for ri in range(3, ws.max_row):  # 薪資表從第3列跑到第N列
    value = ws.cell(ri, 3).value  # 取得人名
    if value in config:
        print("開始拆分 " + value)
        employee = config[value]
        code = value
        name = employee['name']
        pwd = employee['id']
        email = employee['email']
        wb2 = openpyxl.load_workbook(input_file)  # 開啟副本
        ws2 = wb2.worksheets[0]  # 取第一張表
        ws2.move_range("A" + str(ri) + ":AZ" + str(ri), rows=-(ri-3),
                       cols=0, translate=True)  # 移動目標列到第3列
        ws2.delete_rows(4, ws.max_row)  # 刪除多餘列
        output_file = os.path.join(tmp_dir, code + '.xlsx')
        wb2.save(output_file)
        tmp_files.append(output_file)
    else:
        break

print('拆分完成!')


# 加密--------------------------------------------------------------------------------
input('是否開始加密?')
for i in range(len(tmp_files)):
    pwd_file = tmp_files[i]
    value = os.path.basename(pwd_file).split('.')[0]
    employee = config[value]
    print('加密 ' + str(value))
    set_password(pwd_file, employee['id'])
print('加密完成!')

# 發送email--------------------------------------------------------------------------------
subject = input('即將開始發送email,請輸入主旨:')
for i in range(len(tmp_files)):
    email_file = tmp_files[i]
    value = os.path.basename(email_file).split('.')[0]
    employee = config[value]
    print('發送email ' + str(value))
    sendExcelByMail(subject, email_file, employee['email'])
shutil.rmtree(tmp_dir)
print('Done!')
